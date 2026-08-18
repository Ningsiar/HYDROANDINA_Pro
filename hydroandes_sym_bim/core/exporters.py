# -*- coding: utf-8 -*-
"""
core/exporters.py

Exportación de resultados a los formatos solicitados, garantizando que
los mismos valores numéricos se reflejen en todos los archivos (se
exporta siempre desde las mismas estructuras de datos en memoria, nunca
se recalcula nada por formato de salida).
"""
import csv
import os
from typing import List, Dict

from qgis.core import (
    QgsVectorFileWriter, QgsCoordinateTransformContext, QgsRasterFileWriter,
    QgsRasterPipe,
)


def exportar_vector(layer, ruta_base_sin_extension: str, crs=None):
    """
    Exporta una capa vectorial (polígono de cuenca o red de drenaje) a
    .shp, .kml y .geojson, todos con el mismo contenido geométrico y de
    atributos.
    Devuelve dict {'shp': ruta, 'kml': ruta, 'geojson': ruta}.
    """
    salidas = {}
    formatos = {
        "shp": ("ESRI Shapefile", ".shp"),
        "kml": ("KML", ".kml"),
        "geojson": ("GeoJSON", ".geojson"),
    }
    transform_context = QgsCoordinateTransformContext()

    for clave, (driver, ext) in formatos.items():
        ruta = ruta_base_sin_extension + ext
        opciones = QgsVectorFileWriter.SaveVectorOptions()
        opciones.driverName = driver
        opciones.fileEncoding = "UTF-8"
        if crs is not None:
            opciones.ct = None  # se asume que layer ya está en el CRS deseado;
            # si se requiere reproyectar, hacerlo antes con native:reprojectlayer.
        resultado = QgsVectorFileWriter.writeAsVectorFormatV3(
            layer, ruta, transform_context, opciones
        )
        if resultado[0] != QgsVectorFileWriter.NoError:
            raise RuntimeError(f"Error exportando a {clave}: {resultado[1]}")
        salidas[clave] = ruta

    return salidas


def exportar_raster_tif(raster_layer, ruta_tif: str):
    """Exporta una QgsRasterLayer (p.ej. el MDE recortado a la cuenca) a GeoTIFF."""
    provider = raster_layer.dataProvider()
    pipe = QgsRasterPipe()
    if not pipe.set(provider.clone()):
        raise RuntimeError("No se pudo preparar el pipeline de exportación ráster.")
    writer = QgsRasterFileWriter(ruta_tif)
    writer.writeRaster(
        pipe, provider.xSize(), provider.ySize(), provider.extent(), raster_layer.crs()
    )
    return ruta_tif


def exportar_tablas_csv(tablas: Dict[str, List[dict]], carpeta_destino: str) -> List[str]:
    """
    tablas: {'grupo1_basicos': [ {...}, ... ], 'tc_methods': [...], 'curve_number': [...] }
    Escribe un .csv por cada tabla. Devuelve la lista de rutas generadas.
    """
    rutas = []
    os.makedirs(carpeta_destino, exist_ok=True)
    for nombre, filas in tablas.items():
        if not filas:
            continue
        ruta = os.path.join(carpeta_destino, f"{nombre}.csv")
        with open(ruta, "w", newline="", encoding="utf-8") as f:
            campos = list(filas[0].keys())
            writer = csv.DictWriter(f, fieldnames=campos)
            writer.writeheader()
            for fila in filas:
                writer.writerow(fila)
        rutas.append(ruta)
    return rutas


def exportar_tablas_excel(tablas: Dict[str, List[dict]], ruta_xlsx: str):
    """
    Requiere openpyxl (no forma parte de qgis.core; si el usuario no lo
    tiene instalado en el intérprete de Python de QGIS, se debe instalar
    con el OSGeo4W Shell: `pip install openpyxl`). Se importa de forma
    perezosa para no romper el resto del plugin si falta la librería.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl no está instalado en el intérprete de Python de QGIS. "
            "Ejecute 'pip install openpyxl' desde el OSGeo4W Shell (Windows) "
            "o el terminal con el Python de QGIS (Linux/Mac)."
        ) from e

    wb = openpyxl.Workbook()
    primera = True
    for nombre_hoja, filas in tablas.items():
        if not filas:
            continue
        ws = wb.active if primera else wb.create_sheet()
        ws.title = nombre_hoja[:31]  # límite de Excel para nombres de hoja
        primera = False

        campos = list(filas[0].keys())
        for col_idx, campo in enumerate(campos, start=1):
            ws.cell(row=1, column=col_idx, value=campo)
        for row_idx, fila in enumerate(filas, start=2):
            for col_idx, campo in enumerate(campos, start=1):
                ws.cell(row=row_idx, column=col_idx, value=fila.get(campo))
        for col_idx, campo in enumerate(campos, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(campo) + 2)

    wb.save(ruta_xlsx)
    return ruta_xlsx
