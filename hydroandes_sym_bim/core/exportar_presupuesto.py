# -*- coding: utf-8 -*-
"""
core/exportar_presupuesto.py

Exportación de un core.presupuesto.Presupuesto a Excel con DOS
plantillas distintas:

  1. `exportar_s10_xlsx()` -- estructura de 3 hojas (Presupuesto,
     Análisis de Precios Unitarios, Relación de Insumos) que reproduce
     el diseño de un reporte REAL de S10 Presupuestos -- el layout se
     aprendió al parsear los PDF reales que el usuario aportó en la
     Fase 4 del Módulo Presupuesto (proyecto de referencia Cajamarca,
     ver core/apu_referencia.py): bloque de encabezado con datos de la
     obra, partidas agrupadas por Título/Subtítulo con subtotal por
     grupo, y el desglose de APU por categoría (Mano de Obra /
     Materiales / Equipos / Herramienta Manual / Subcontratos /
     Subpartida) con subtotal por categoría. Los NOMBRES de columna y
     el orden son los mismos que aparecen en esos reportes reales --
     NO es un archivo importable de vuelta a S10 (S10 usa un formato
     de base de datos propietario, .s10/.rpu, no Excel), es un reporte
     de SALIDA con la misma apariencia, útil para entregar al cliente
     o para copiar/pegar hacia un expediente que sí use S10.

  2. `exportar_generico_xlsx()` -- una plantilla tabular GENÉRICA (2
     hojas: Presupuesto e Insumos, columnas simples sin agrupación),
     pensada como punto de partida para importar a OTRO software de
     presupuestos (p.ej. Delphin Xpress) que acepte datos tabulares.
     ADVERTENCIA EXPLÍCITA: no existe un archivo de muestra real de
     Delphin Xpress con el que verificar el formato exacto que ese
     software espera (a diferencia de S10, donde sí se tuvo un PDF
     real) -- este exportador NO está verificado contra una
     importación real a Delphin Xpress. Si cuenta con un archivo de
     exportación real de ese software, se puede ajustar esta función
     para que coincida exactamente.

openpyxl se importa de forma perezosa, igual que en
core/table_export.py, para no romper el resto del plugin si falta en
el intérprete de Python de QGIS.
"""
import os

from . import presupuesto as pres


class ExportarPresupuestoError(Exception):
    pass


def _importar_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side
        return openpyxl, Font, Border, Side
    except ImportError as e:
        raise ExportarPresupuestoError(
            "openpyxl no está instalado en el intérprete de Python de QGIS. Ejecute "
            "'pip install openpyxl' desde el OSGeo4W Shell (Windows) o el terminal con el "
            "Python de QGIS (Linux/Mac), y vuelva a intentar."
        ) from e


def _autoajustar_columnas(ws, anchos_minimos=None):
    anchos_minimos = anchos_minimos or {}
    for columna in ws.columns:
        letra = None
        ancho = 8
        for celda in columna:
            if celda.value is None:
                continue
            letra = celda.column_letter
            ancho = max(ancho, len(str(celda.value)) + 2)
        if letra:
            ancho = max(ancho, anchos_minimos.get(letra, 0))
            ws.column_dimensions[letra].width = min(ancho, 55)


def _agrupar_partidas_por_grupo(partidas):
    """Agrupa manteniendo el orden de aparición -- {grupo: [Partida,...]}.
    `grupo=""` (sin Título/Subtítulo asignado) se conserva como su
    propio grupo, con ese texto vacío -- no se descarta ninguna partida."""
    grupos = {}
    orden = []
    for p in partidas:
        if p.grupo not in grupos:
            grupos[p.grupo] = []
            orden.append(p.grupo)
        grupos[p.grupo].append(p)
    return [(g, grupos[g]) for g in orden]


def exportar_s10_xlsx(presupuesto: "pres.Presupuesto", ruta: str, datos_obra: dict = None) -> str:
    """Escribe un .xlsx de 3 hojas con el layout de un reporte real de
    S10 Presupuestos -- ver docstring del módulo. `datos_obra` (todo
    opcional): {"obra":.., "cliente":.., "lugar":.., "fecha":..,
    "moneda":.., "plazo":..} -- se muestran en el bloque de encabezado
    de la hoja "Presupuesto" si se indican."""
    openpyxl, Font, Border, Side = _importar_openpyxl()
    datos_obra = datos_obra or {}
    resumen = presupuesto.resumen()
    negrita = Font(bold=True)
    negrita_grande = Font(bold=True, size=12)
    borde_inferior = Border(bottom=Side(style="thin"))

    wb = openpyxl.Workbook()

    # ======================================================================
    # Hoja 1: PRESUPUESTO
    # ======================================================================
    ws1 = wb.active
    ws1.title = "Presupuesto"
    ws1.append(["PRESUPUESTO DE OBRA"])
    ws1["A1"].font = negrita_grande
    fila = 2
    for etiqueta, clave in (("Obra", "obra"), ("Cliente", "cliente"), ("Lugar", "lugar"),
                            ("Fecha", "fecha"), ("Plazo de ejecución", "plazo")):
        if datos_obra.get(clave):
            ws1.append([f"{etiqueta}:", datos_obra[clave]])
            fila += 1
    ws1.append([f"Moneda: {datos_obra.get('moneda', 'Soles (S/.)')}"])
    fila += 1
    ws1.append([])
    fila += 1
    fila_encabezado = fila + 1
    ws1.append(["Item", "Descripción", "Und.", "Metrado", "Precio S/.", "Parcial S/."])
    for celda in ws1[fila_encabezado]:
        celda.font = negrita
        celda.border = borde_inferior
    fila = fila_encabezado + 1

    for grupo, partidas_grupo in _agrupar_partidas_por_grupo(presupuesto.partidas):
        if grupo:
            subtotal_grupo = sum(p.costo_parcial() for p in partidas_grupo)
            ws1.append([None, grupo, None, None, None, round(subtotal_grupo, 2)])
            for celda in ws1[fila]:
                celda.font = negrita
            fila += 1
        for p in partidas_grupo:
            ws1.append([p.codigo, p.descripcion, p.unidad, p.metrado,
                        round(p.precio_unitario(), 2), round(p.costo_parcial(), 2)])
            fila += 1

    ws1.append([])
    fila += 1
    filas_resumen = [
        ("COSTO DIRECTO", resumen["costo_directo"]),
        (f"GASTOS GENERALES ({resumen['gastos_generales_pct']:g}%)", resumen["gastos_generales"]),
        (f"UTILIDAD ({resumen['utilidad_pct']:g}%)", resumen["utilidad"]),
        ("SUBTOTAL", resumen["subtotal"]),
        (f"IGV ({resumen['igv_pct']:g}%)", resumen["igv"]),
        ("PRESUPUESTO TOTAL", resumen["total"]),
    ]
    for etiqueta, valor in filas_resumen:
        ws1.append([None, None, None, None, etiqueta, valor])
        for celda in ws1[fila]:
            if celda.value is not None:
                celda.font = negrita
        fila += 1
    _autoajustar_columnas(ws1, {"B": 45})

    # ======================================================================
    # Hoja 2: ANÁLISIS DE PRECIOS UNITARIOS
    # ======================================================================
    ws2 = wb.create_sheet("Análisis de Precios Unitarios")
    fila = 1
    for grupo, partidas_grupo in _agrupar_partidas_por_grupo(presupuesto.partidas):
        for p in partidas_grupo:
            if p.apu is None:
                continue
            ws2.cell(row=fila, column=1, value=f"Partida: {p.codigo}  {p.descripcion}").font = negrita_grande
            fila += 1
            ws2.cell(row=fila, column=1, value=f"Rendimiento:  {p.unidad}/DIA")
            ws2.cell(row=fila, column=4, value=f"Costo unitario directo por: {p.unidad}")
            ws2.cell(row=fila, column=6, value=round(p.precio_unitario(), 2)).font = negrita
            fila += 1
            fila_encab = fila
            ws2.append(["Descripción Recurso", "Unidad", "Cuadrilla", "Cantidad", "Precio S/.", "Parcial S/."])
            for celda in ws2[fila_encab]:
                celda.font = negrita
                celda.border = borde_inferior
            fila += 1
            subtotales = p.apu.subtotal_por_tipo()
            for tipo in pres.TipoInsumo.TODOS:
                items_tipo = [i for i in p.apu.items if i.insumo.tipo == tipo]
                if not items_tipo:
                    continue
                ws2.cell(row=fila, column=1, value=tipo).font = negrita
                fila += 1
                for item in items_tipo:
                    usa_rendimiento = item.rendimiento is not None
                    ws2.append([
                        item.insumo.descripcion, item.insumo.unidad,
                        round(item.cuadrilla, 4) if usa_rendimiento and item.cuadrilla is not None else None,
                        round(item.cantidad_por_unidad(), 4),
                        round(item.insumo.precio_unitario, 2), round(item.parcial_por_unidad(), 4),
                    ])
                    fila += 1
                ws2.cell(row=fila, column=1, value=f"Subtotal {tipo}").font = negrita
                ws2.cell(row=fila, column=6, value=round(subtotales[tipo], 4)).font = negrita
                fila += 1
            ws2.append([])
            fila += 1
    _autoajustar_columnas(ws2, {"A": 40})

    # ======================================================================
    # Hoja 3: RELACIÓN DE INSUMOS
    # ======================================================================
    ws3 = wb.create_sheet("Relación de Insumos")
    ws3.append(["Código", "Recurso", "Unidad", "Cantidad", "Precio S/.", "Parcial S/."])
    for celda in ws3[1]:
        celda.font = negrita
        celda.border = borde_inferior
    fila = 2
    relacion = presupuesto.relacion_insumos()
    tipo_actual = None
    costo_total_general = 0.0
    for f in relacion:
        if f["tipo"] != tipo_actual:
            tipo_actual = f["tipo"]
            ws3.cell(row=fila, column=1, value=tipo_actual).font = negrita
            fila += 1
        ws3.append([f["codigo"], f["descripcion"], f["unidad"], f["cantidad_total"],
                    f["precio_unitario"], f["costo_total"]])
        costo_total_general += f["costo_total"]
        fila += 1
    ws3.append([])
    fila += 1
    ws3.cell(row=fila, column=2, value="TOTAL").font = negrita
    ws3.cell(row=fila, column=6, value=round(costo_total_general, 2)).font = negrita
    _autoajustar_columnas(ws3, {"B": 40})

    os.makedirs(os.path.dirname(ruta) or ".", exist_ok=True)
    wb.save(ruta)
    return ruta


def exportar_generico_xlsx(presupuesto: "pres.Presupuesto", ruta: str) -> str:
    """Plantilla tabular GENÉRICA (2 hojas simples, sin agrupación ni
    bloques de encabezado) -- punto de partida para importar a otro
    software de presupuestos (p.ej. Delphin Xpress). NO verificada
    contra un archivo de importación real de ese software -- ver
    advertencia completa en el docstring del módulo."""
    openpyxl, Font, Border, Side = _importar_openpyxl()
    negrita = Font(bold=True)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Presupuesto"
    ws1.append(["Codigo", "Descripcion", "Unidad", "Metrado", "PrecioUnitario", "Parcial", "Grupo"])
    for celda in ws1[1]:
        celda.font = negrita
    for p in presupuesto.partidas:
        ws1.append([p.codigo, p.descripcion, p.unidad, p.metrado,
                    round(p.precio_unitario(), 2), round(p.costo_parcial(), 2), p.grupo])
    _autoajustar_columnas(ws1, {"B": 45, "G": 30})

    ws2 = wb.create_sheet("Insumos")
    ws2.append(["Codigo", "Descripcion", "Unidad", "Tipo", "Cantidad", "PrecioUnitario", "CostoTotal"])
    for celda in ws2[1]:
        celda.font = negrita
    for f in presupuesto.relacion_insumos():
        ws2.append([f["codigo"], f["descripcion"], f["unidad"], f["tipo"],
                    f["cantidad_total"], f["precio_unitario"], f["costo_total"]])
    _autoajustar_columnas(ws2, {"B": 40})

    os.makedirs(os.path.dirname(ruta) or ".", exist_ok=True)
    wb.save(ruta)
    return ruta
